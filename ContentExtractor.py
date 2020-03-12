"""
Created on unknown

@author: soupnook
"""

import os
import openpyxl
from TableExtractor import *
from MetaDataHandler import *
from RasaTagger import *
from StanfordNERtagger import *
from TableExtractor import *

class ContentExtractor:
    '''
    Extract sov data from excel files using openpyxl library
    '''
    def ExcelExtractor(self,excel_files,outputpath):
        for eachfile in excel_files:

            #path for output file
            req_output_filename = eachfile[eachfile.find('\\') + 1:].replace(".xlsx","_results.xlsx")
            outputfilepath = outputpath+req_output_filename

            if ("~" not in eachfile):
                print("File : "+eachfile)
                workbook = openpyxl.load_workbook(filename=eachfile, data_only=True)

                #checking if multiple sheets are there or not
                if(len(workbook.worksheets)>1):
                    MetaDataHandler.multiple_sheets=True

                #Handling Merge cells in multiple sheets
                for sheet in workbook.worksheets:
                    sheet,merge_flag = TableExtractor().merge_cell_handler(sheet)
                    if(merge_flag == 1):
                        MetaDataHandler.merge_cells = True
                    #print(workbook.sheetnames)

                sov_sheet = []
                non_sov_sheet = []
                req_sheetname = ""
                table_flag = 0
                #Incase of multiple sheets, logic to fetch the required sheet
                if(len(workbook.sheetnames)>1):
                    #find if any sheet have sov as substring in sheet name
                    #if not then based on header row of sheet, getting required sov sheet
                    for sheetname in workbook.sheetnames:
                        if("sov" in sheetname.lower().replace(" ","")):
                            sov_sheet.append(sheetname)
                        else:
                            non_sov_sheet.append(sheetname)

                    if(len(sov_sheet)>0):
                        if(len(sov_sheet)==1):
                            req_sheetname = sov_sheet[0]
                        else:
                            req_sheetname,exact_table = TableExtractor().get_required_sheet(sov_sheet,workbook)
                            table_flag = 1
                    elif(len(non_sov_sheet)>0):
                        if(len(non_sov_sheet)==1):
                            req_sheetname = non_sov_sheet[0]
                        else:
                            req_sheetname,exact_table = TableExtractor().get_required_sheet(non_sov_sheet, workbook)
                            table_flag = 1
                    else:
                        req_sheetname = workbook.active
                else:
                    req_sheetname = (workbook.sheetnames[0])

            #getting the required sheet name
            print("required sheetname: ",req_sheetname)

            #if there is only one sheet
            if(table_flag == 0):
                sheetdata = workbook[req_sheetname]
                table = TableExtractor().extract_full_table(sheetdata)
                req_table = TableExtractor().extract_exact_table(table)
            else:
                req_table = exact_table

            ##training the model for header mapping
            Results().create_outputfile(outputfilepath)
            RasaTagger().Tag_headers_Rasa(req_table,outputfilepath)
            StanfordNERtagger().Tag_headers_stanford(req_table,outputfilepath)
            print("-----" * 25)

    ##Extract sov data from excel files using Abbyy Fine reader
    def PDFExtractor(self,pdf_files,outputpath):
        for eachfile in pdf_files:
            print(eachfile)
            #abbyyExtractor(eachfile)
